"""
Carga del PUC completo de la empresa desde el archivo que exporta su software
contable (retroalimentación del contador: cada empresa —y cada sector: real,
solidario, financiero, seguros…— tiene su propio plan de cuentas, y hay que
poder subirlo para causar en la AUXILIAR correcta, no en la cuenta mayor).

Acepta CSV (separador ; o ,) y Excel (.xlsx). Autodetecta las columnas de
código y nombre por su encabezado (o, si no hay encabezado reconocible, toma
la primera columna con pinta de código y la de al lado como nombre). Es
REENTRANTE: vuelve a subir el mismo archivo y solo actualiza nombres, sin
duplicar (clave: código por empresa).
"""
import csv
import io
import re
from dataclasses import dataclass, field

ENCABEZADOS_CODIGO = {"codigo", "código", "cuenta", "cuenta contable", "code",
                      "cta", "nro cuenta", "numero cuenta", "número cuenta"}
ENCABEZADOS_NOMBRE = {"nombre", "descripcion", "descripción", "nombre cuenta",
                      "nombre de cuenta", "detalle", "concepto", "name", "cuenta nombre"}

_SOLO_CODIGO = re.compile(r"^\d[\d.\-]{0,19}$")  # 5110, 5110.35, 51-10-35…


class PUCInvalido(Exception):
    """El archivo no se pudo leer como PUC. Mensaje apto para el usuario."""


@dataclass
class ResumenPUC:
    creadas: int = 0
    actualizadas: int = 0
    sin_cambio: int = 0
    ignoradas: int = 0
    total_archivo: int = 0
    errores: list = field(default_factory=list)

    @property
    def procesadas(self):
        return self.creadas + self.actualizadas + self.sin_cambio


def _norm(texto):
    return (texto or "").strip().strip('"').lower()


def _limpiar_codigo(valor):
    """Deja solo dígitos (el PUC se compara por dígitos: 5110.35 == 511035)."""
    return re.sub(r"\D", "", (valor or "").strip())


def _filas_de_csv(contenido):
    texto = contenido.decode("utf-8-sig", errors="replace")
    muestra = texto[:2048]
    sep = ";" if muestra.count(";") >= muestra.count(",") else ","
    return list(csv.reader(io.StringIO(texto), delimiter=sep))


def _filas_de_xlsx(contenido):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise PUCInvalido("no se pudo leer el Excel en este servidor; súbelo como CSV.")
    try:
        libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception:
        raise PUCInvalido("el archivo no es un Excel (.xlsx) válido.")
    hoja = libro.active
    filas = []
    for fila in hoja.iter_rows(values_only=True):
        filas.append(["" if celda is None else str(celda) for celda in fila])
    libro.close()
    return filas


def _detectar_columnas(filas):
    """→ (idx_codigo, idx_nombre, fila_inicio). Usa el encabezado si lo reconoce;
    si no, deduce por el contenido (primera columna que parece código)."""
    for i, fila in enumerate(filas[:5]):
        normalizados = [_norm(c) for c in fila]
        idx_cod = next((j for j, c in enumerate(normalizados)
                        if c in ENCABEZADOS_CODIGO), None)
        idx_nom = next((j for j, c in enumerate(normalizados)
                        if c in ENCABEZADOS_NOMBRE), None)
        if idx_cod is not None and idx_nom is not None:
            return idx_cod, idx_nom, i + 1

    # Sin encabezado reconocible: buscar una fila de datos con un código y texto.
    for fila in filas:
        for j, celda in enumerate(fila):
            if _SOLO_CODIGO.match((celda or "").strip()):
                # el nombre es la siguiente columna no vacía
                for k in range(j + 1, len(fila)):
                    if (fila[k] or "").strip():
                        return j, k, 0
    raise PUCInvalido(
        "no se reconocieron columnas de código y nombre. Usa un archivo con "
        "encabezados «Código» y «Nombre», o dos columnas: el código y el nombre.")


def leer_filas_puc(nombre_archivo, contenido):
    """→ lista de (codigo_digitos, nombre) desde CSV o XLSX, ya limpia."""
    bajo = (nombre_archivo or "").lower()
    if bajo.endswith((".xlsx", ".xlsm")):
        filas = _filas_de_xlsx(contenido)
    else:
        filas = _filas_de_csv(contenido)
    filas = [f for f in filas if any((c or "").strip() for c in f)]
    if not filas:
        raise PUCInvalido("el archivo está vacío.")

    idx_cod, idx_nom, inicio = _detectar_columnas(filas)
    cuentas = []
    for fila in filas[inicio:]:
        if idx_cod >= len(fila):
            continue
        codigo = _limpiar_codigo(fila[idx_cod])
        nombre = (fila[idx_nom].strip() if idx_nom < len(fila) else "").strip('"')
        if not codigo:
            continue
        cuentas.append((codigo, nombre or f"Cuenta {codigo}"))
    if not cuentas:
        raise PUCInvalido("no se encontró ninguna cuenta con código en el archivo.")
    return cuentas


def importar_puc(empresa, nombre_archivo, contenido, reemplazar=False):
    """Carga/actualiza el catálogo PUC de la empresa. Reentrante por código.

    reemplazar=True borra el catálogo previo antes de cargar (para reemplazar
    un PUC completo por otro). → ResumenPUC."""
    from .models import CuentaPUC

    cuentas = leer_filas_puc(nombre_archivo, contenido)
    resumen = ResumenPUC(total_archivo=len(cuentas))

    if reemplazar:
        CuentaPUC.objects.de_empresa(empresa).delete()

    existentes = {c.codigo: c for c in CuentaPUC.objects.de_empresa(empresa)}
    vistos = set()
    nuevas = []
    for codigo, nombre in cuentas:
        nombre = nombre[:200]
        if codigo in vistos:  # duplicado dentro del propio archivo
            resumen.ignoradas += 1
            continue
        vistos.add(codigo)
        actual = existentes.get(codigo)
        if actual is None:
            nuevas.append(CuentaPUC(empresa=empresa, codigo=codigo, nombre=nombre))
            resumen.creadas += 1
        elif actual.nombre != nombre:
            actual.nombre = nombre
            actual.save(update_fields=["nombre"])
            resumen.actualizadas += 1
        else:
            resumen.sin_cambio += 1
    if nuevas:
        CuentaPUC.objects.bulk_create(nuevas)
    return resumen
